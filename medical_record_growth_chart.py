# -*- coding: UTF-8 -*-
from PyQt5 import QtWidgets, QtCore, QtChart, QtGui
import openpyxl
import os
import json

from libs import class_utils
from libs import ui_utils
from libs import string_utils
from libs import system_utils
from libs import number_utils
from libs import date_utils

CURRENT_DIR = os.path.abspath(os.path.join(os.path.dirname("__file__")))


# 病歷資料-兒童生長曲線 2022.06.04
class MedicalRecordGrowthChart(QtWidgets.QMainWindow):
    # 初始化
    def __init__(self, parent=None, *args):
        super(MedicalRecordGrowthChart, self).__init__(parent)
        self.parent = parent
        self.database = args[0]
        self.system_settings = args[1]
        self.case_key = args[2]
        self.call_from = args[3]

        self.treat_type = None
        self.doctor_done = False
        self.ui = None

        self.age_list = [
            0.5, 1, 1.5, 2, 2.5, 3, 3.5, 4, 4.5, 5, 5.5, 6, 6.5,
            7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17
        ]
        self._set_ui()
        self._set_signal()
        self._set_standard_list()

        self.user_name = system_utils.get_user_name(self.system_settings)

    # 解構
    def __del__(self):
        self.close_all()

    # 關閉
    def close_all(self):
        pass

    # 設定GUI
    def _set_ui(self):
        self.ui = ui_utils.load_ui_file(ui_utils.UI_MEDICAL_RECORD_GROWTH_CHART, self)
        system_utils.set_css(self, self.system_settings)
        self.table_widget_growth = class_utils.get_table_widget(self.ui.tableWidget_growth, self.database)
        self._set_table_width()

    # 設定欄位寬度
    def _set_table_width(self):
        width = [60, 150, 80, 110, 80, 110]
        self.table_widget_growth.set_table_heading_width(width)

    # 設定信號
    def _set_signal(self):
        pass

    def plot_chart(self):
        self._read_data()

        try:
            # self._read_data()
            self._plot_chart()
        except Exception:
            pass

    def _set_table_widget(self):
        self.ui.tableWidget_growth.setRowCount(0)  # clear data

        self.ui.tableWidget_growth.setRowCount(len(self.age_list))
        for row_no, age in enumerate(self.age_list):
            self.ui.tableWidget_growth.setItem(
                row_no, 0, QtWidgets.QTableWidgetItem(str(age))
            )
            self.ui.tableWidget_growth.item(
                row_no, 0).setTextAlignment(
                QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter
            )

    def _read_data(self):
        self._set_table_widget()

        patient_key = self.parent.patient_record['PatientKey']
        if patient_key is None:
            return

        sql = f'''
            SELECT case_extension.*, cases.CaseDate FROM case_extension
                LEFT JOIN cases ON case_extension.CaseKey = cases.CaseKey
            WHERE
                PatientKey = {patient_key} AND
                ExtensionType = "診前檢查"
            ORDER BY cases.CaseDate DESC
        '''
        rows = self.database.select_record(sql)
        if len(rows) <= 0:
            return

        for row in rows:
            self._set_data(row)

        self.ui.tableWidget_growth.resizeRowsToContents()

    def _get_row_no(self, age):
        for row_no in range(self.ui.tableWidget_growth.rowCount()):
            row_age = number_utils.get_float(self.ui.tableWidget_growth.item(row_no, 0).text())
            if age == row_age:
                return row_no

        return None

    def _get_percent(self, age, value, percent_list):
        row_no = self._get_row_no(age)
        current_list = percent_list[row_no+1]

        value_index = 0
        for index, standard_list in enumerate(current_list):
            if number_utils.get_float(value) < standard_list:
                value_index = index
                break

        if value_index == 0:
            current_percent = '3rd'
        elif value_index == 1:
            current_percent = '15th'
        elif value_index == 2:
            current_percent = '25th'
        elif value_index == 3:
            current_percent = '50th'
        elif value_index == 4:
            current_percent = '75th'
        elif value_index == 5:
            current_percent = '85th'
        elif value_index == 6:
            current_percent = '97th'
        else:
            current_percent = 'N/A'

        return current_percent
    
    def _set_data(self, row):
        birthday = self.parent.patient_record['Birthday']
        case_date = row['CaseDate'].date()
        year, month = date_utils.get_age(birthday, case_date)
        age = year
        if year <= 6 and month > 6:
            age += 0.5

        # if self.ui.tableWidget_growth.item(age_month, 1) is not None:  # 有資料就不要再次顯示
        #     return

        row_no = self._get_row_no(age)
        if row_no is None:
            return

        exam_dict = json.loads(string_utils.xstr(row['Content']))
        height = exam_dict['height']
        weight = exam_dict['weight']
        height_percent = self._get_percent(age, height, self.height_list)
        weight_percent = self._get_percent(age, weight, self.weight_list)

        case_date_str = case_date.strftime('%Y-%m-%d')
        case_date_item = self.ui.tableWidget_growth.item(row_no, 1)
        if case_date_item is not None:
            case_date_str += f'\n{case_date_item.text()}'

        height_item = self.ui.tableWidget_growth.item(row_no, 2)
        if height_item is not None:
            height += f'\n{height_item.text()}'

        height_percent_item = self.ui.tableWidget_growth.item(row_no, 3)
        if height_percent_item is not None:
            height_percent += f'\n{height_percent_item.text()}'

        weight_item = self.ui.tableWidget_growth.item(row_no, 4)
        if weight_item is not None:
            weight += f'\n{weight_item.text()}'

        weight_percent_item = self.ui.tableWidget_growth.item(row_no, 5)
        if weight_percent_item is not None:
            weight_percent += f'\n{weight_percent_item.text()}'

        growth_data = [
            case_date_str,
            height,
            height_percent,
            weight,
            weight_percent,
        ]

        for col_no, data in enumerate(growth_data):
            self.ui.tableWidget_growth.setItem(
                row_no, col_no+1, QtWidgets.QTableWidgetItem(data)
            )
            if col_no in [0, 2, 4]:
                self.ui.tableWidget_growth.item(
                    row_no, col_no+1).setTextAlignment(
                    QtCore.Qt.AlignCenter | QtCore.Qt.AlignVCenter
                )
            elif col_no in [1, 3]:
                self.ui.tableWidget_growth.item(
                    row_no, col_no+1).setTextAlignment(
                    QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter
                )

    def _plot_chart(self):
        while self.ui.verticalLayout_chart.count():
            item = self.ui.verticalLayout_chart.takeAt(0)
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()

        self._plot_growth_chart()

    def _plot_growth_chart(self):
        chart = QtChart.QChart()
        chart.setTitle(f'{self.gender}孩生長曲線圖')
        chart.setAnimationOptions(QtChart.QChart.SeriesAnimations)

        self._plot_standard_lines(chart)

        self._plot_growth_height(chart)
        self._plot_growth_weight(chart)

        # x_points = [0.5, 1, 1.5, 2, 2.5, 3, 3.5, 4, 4.5, 5, 5.5, 6, 6.5, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17]
        # x_axis = QtChart.QCategoryAxis()
        # for value, x in enumerate(x_points):
        #     x_axis.append(str(x), value)

        # x_axis.setTickCount(24)

        # y_axis = QtChart.QValueAxis()
        # y_axis.setRange(40.0, 180.0)
        # y_axis.setLabelFormat('%0.1f')
        # y_axis.setTickCount(20)

        # chart.setAxisX(x_axis)
        # chart.setAxisY(y_axis)
        chart.createDefaultAxes()

        chart.legend().setVisible(False)
        chart.legend().setAlignment(QtCore.Qt.AlignRight)

        self.chartView = QtChart.QChartView(chart)
        self.chartView.setRenderHint(QtGui.QPainter.Antialiasing)

        self.ui.verticalLayout_chart.addWidget(self.chartView)

    def _set_standard_list(self):
        self.gender = string_utils.xstr(self.parent.patient_record['Gender'])
        if self.gender not in ['男', '女']:
            self.gender = '男'

        self.height_list = []
        self.weight_list = []

        sub_dir = 'tables'
        standard_file = os.path.join(CURRENT_DIR, sub_dir, '0-17.xlsx')

        wb = openpyxl.load_workbook(standard_file)
        if self.gender == '男':
            sheet_name = '男孩'
        else:
            sheet_name = '女孩'

        sheet = wb[sheet_name]

        start_no = 4
        for row_no in range(start_no, len(self.age_list)+start_no+1):
            self.height_list.append([
                number_utils.get_float(sheet.cell(row_no, 10).value),
                number_utils.get_float(sheet.cell(row_no, 11).value),
                number_utils.get_float(sheet.cell(row_no, 12).value),
                number_utils.get_float(sheet.cell(row_no, 13).value),
                number_utils.get_float(sheet.cell(row_no, 14).value),
                number_utils.get_float(sheet.cell(row_no, 15).value),
                number_utils.get_float(sheet.cell(row_no, 16).value),
            ])

            self.weight_list.append([
                number_utils.get_float(sheet.cell(row_no, 2).value),
                number_utils.get_float(sheet.cell(row_no, 3).value),
                number_utils.get_float(sheet.cell(row_no, 4).value),
                number_utils.get_float(sheet.cell(row_no, 5).value),
                number_utils.get_float(sheet.cell(row_no, 6).value),
                number_utils.get_float(sheet.cell(row_no, 7).value),
                number_utils.get_float(sheet.cell(row_no, 8).value),
            ])

        wb.close()

    def _plot_standard_lines(self, chart):
        series_height_head3 = QtChart.QLineSeries(name='身高P3')
        series_height_head15 = QtChart.QLineSeries(name='身高P15')
        series_height_head25 = QtChart.QLineSeries(name='身高P25')
        series_height_head50 = QtChart.QLineSeries(name='身高P50')
        series_height_head75 = QtChart.QLineSeries(name='身高P75')
        series_height_head85 = QtChart.QLineSeries(name='身高P85')
        series_height_head97 = QtChart.QLineSeries(name='身高P97')

        # series_height_head97.setPointLabelsVisible(True)

        for i in range(len(self.height_list)):
            series_height_head3.append(i, self.height_list[i][0])
            series_height_head15.append(i, self.height_list[i][1])
            series_height_head25.append(i, self.height_list[i][2])
            series_height_head50.append(i, self.height_list[i][3])
            series_height_head75.append(i, self.height_list[i][4])
            series_height_head85.append(i, self.height_list[i][5])
            series_height_head97.append(i, self.height_list[i][6])

        chart.addSeries(series_height_head97)
        chart.addSeries(series_height_head85)
        chart.addSeries(series_height_head75)
        chart.addSeries(series_height_head50)
        chart.addSeries(series_height_head25)
        chart.addSeries(series_height_head15)
        chart.addSeries(series_height_head3)

        series_weight_head3 = QtChart.QLineSeries(name='體重P3')
        series_weight_head15 = QtChart.QLineSeries(name='體重P15')
        series_weight_head25 = QtChart.QLineSeries(name='體重P25')
        series_weight_head50 = QtChart.QLineSeries(name='體重P50')
        series_weight_head75 = QtChart.QLineSeries(name='體重P75')
        series_weight_head85 = QtChart.QLineSeries(name='體重P85')
        series_weight_head97 = QtChart.QLineSeries(name='體重P97')

        for i in range(len(self.weight_list)):
            series_weight_head3.append(i, self.weight_list[i][0])
            series_weight_head15.append(i, self.weight_list[i][1])
            series_weight_head25.append(i, self.weight_list[i][2])
            series_weight_head50.append(i, self.weight_list[i][3])
            series_weight_head75.append(i, self.weight_list[i][4])
            series_weight_head85.append(i, self.weight_list[i][5])
            series_weight_head97.append(i, self.weight_list[i][6])

        chart.addSeries(series_weight_head97)
        chart.addSeries(series_weight_head85)
        chart.addSeries(series_weight_head75)
        chart.addSeries(series_weight_head50)
        chart.addSeries(series_weight_head25)
        chart.addSeries(series_weight_head15)
        chart.addSeries(series_weight_head3)

    def _plot_growth_height(self, chart):
        blue = QtCore.Qt.darkBlue
        brush = QtGui.QBrush(blue)

        series = QtChart.QScatterSeries(name='身高')
        series.setMarkerShape(QtChart.QScatterSeries.MarkerShapeCircle)
        series.setMarkerSize(15.0)
        series.setBrush(brush)
        series.setPointLabelsFormat('@yPoint')
        series.setPointLabelsColor(blue)
        series.setPointLabelsVisible(True)

        for row_no in range(self.ui.tableWidget_growth.rowCount()):
            case_date = self.ui.tableWidget_growth.item(row_no, 1)
            if case_date is None:
                continue

            item = self.ui.tableWidget_growth.item(row_no, 2)
            if item is None:
                continue

            height = number_utils.get_float(item.text().split('\n')[-1])
            series.append(row_no, height)

        chart.addSeries(series)

    def _plot_growth_weight(self, chart):
        red = QtCore.Qt.red
        brush = QtGui.QBrush(red)

        series = QtChart.QScatterSeries(name='體重')
        series.setMarkerShape(QtChart.QScatterSeries.MarkerShapeCircle)
        series.setMarkerSize(15.0)
        series.setBrush(brush)
        series.setPointLabelsFormat('@yPoint')
        series.setPointLabelsColor(red)
        series.setPointLabelsVisible(True)

        for row_no in range(self.ui.tableWidget_growth.rowCount()):
            case_date = self.ui.tableWidget_growth.item(row_no, 1)
            if case_date is None:
                continue

            item = self.ui.tableWidget_growth.item(row_no, 4)
            if item is None:
                continue

            weight = number_utils.get_float(item.text().split('\n')[-1])
            series.append(row_no, weight)

        chart.addSeries(series)
