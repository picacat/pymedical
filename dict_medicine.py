# -*- coding: UTF-8 -*-

from PyQt5 import QtWidgets, QtCore
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QPushButton
import csv
import json
import openpyxl
import re

from libs import system_utils
from libs import ui_utils
from libs import db_utils
from libs import module_utils
from libs import personnel_utils
from libs import string_utils


# 處方詞庫 2019.06.12
class DictMedicine(QtWidgets.QMainWindow):
    # 初始化
    def __init__(self, parent=None, *args):
        super(DictMedicine, self).__init__(parent)
        self.parent = parent
        self.args = args
        self.database = args[0]
        self.system_settings = args[1]
        self.ui = None

        self.user_name = system_utils.get_user_name(self.system_settings)

        self._set_ui()
        self._set_signal()

    # 解構
    def __del__(self):
        self.close_all()

    # 關閉
    def close_all(self):
        pass

    # 設定GUI
    def _set_ui(self):
        self.ui = ui_utils.load_ui_file(ui_utils.UI_DICT_MEDICINE, self)
        system_utils.set_css(self, self.system_settings)
        system_utils.center_window(self)

        self.tab_drug = module_utils.get_dict_drug(self, *self.args)
        self.tab_treat = module_utils.get_dict_treat(self, *self.args)
        self.tab_compound = module_utils.get_dict_compound(self, *self.args)
        self.tab_instruction = module_utils.get_dict_instruction(self, *self.args)

        if personnel_utils.get_permission(
                self.database, '處方資料', '執行處方資料', self.user_name) == 'Y':
            self.ui.tabWidget_medicine.addTab(self.tab_drug, '藥品資料')
            self.ui.tabWidget_medicine.addTab(self.tab_treat, '處置資料')
            self.ui.tabWidget_medicine.addTab(self.tab_compound, '成方資料')
            self.ui.tabWidget_medicine.addTab(self.tab_instruction, '指示及醫囑')
        elif personnel_utils.get_permission(
                self.database, '處方資料', '輸入成方資料', self.user_name) == 'Y':
            self.ui.tabWidget_medicine.addTab(self.tab_compound, '成方資料')

        if personnel_utils.get_permission(self.database, '系統作業', '關閉匯出功能', self.user_name) == 'Y':
            self.ui.action_export_dict_medicine_excel.setEnabled(False)
            self.ui.action_export_dict_medicine_json.setEnabled(False)
            self.ui.action_export_dict_compound_json.setEnabled(False)

    # 設定信號
    def _set_signal(self):
        self.ui.action_close.triggered.connect(self.close_template)
        self.ui.action_export_dict_medicine_excel.triggered.connect(self._export_dict_medicine_excel)
        self.ui.action_export_dict_medicine_json.triggered.connect(self._export_dict_medicine_json)
        self.ui.action_export_dict_compound_json.triggered.connect(self._export_dict_compound_json)
        self.ui.action_import_ins_exam_csv.triggered.connect(self._import_ins_exam_csv)
        self.ui.action_import_self_exam_csv.triggered.connect(self._import_self_exam_csv)
        self.ui.action_import_compound_json.triggered.connect(self._import_compound_json)
        self.ui.action_import_supplier_medicine.triggered.connect(self._import_supplier_medicine)

    def close_tab(self):
        current_tab = self.parent.ui.tabWidget_window.currentIndex()
        self.parent.close_tab(current_tab)

    def close_template(self):
        self.close_all()
        self.close_tab()

    def _export_dict_compound_json(self):
        options = QFileDialog.Options()
        json_file_name, _ = QFileDialog.getSaveFileName(
            self.parent,
            "匯出成方JSON檔案", 'compound.json',
            "json檔案 (*.json)",
            options=options
        )
        if not json_file_name:
            return

        sql = '''
            SELECT * FROM refcompound
            ORDER BY CompoundKey
        '''
        rows = self.database.select_record(sql)

        json_data = db_utils.mysql_to_json(rows)
        text_file = open(json_file_name, "w", encoding='utf8')
        text_file.write(str(json_data))
        text_file.close()

        system_utils.show_message_box(
            QMessageBox.Information,
            'JSON資料匯出完成',
            f'<h3>{json_file_name}匯出完成.</h3>',
            'JSON 檔案格式.'
        )

    def _export_dict_medicine_json(self):
        options = QFileDialog.Options()
        json_file_name, _ = QFileDialog.getSaveFileName(
            self.parent,
            "匯出處方JSON檔案", 'medicine.json',
            "json檔案 (*.json)",
            options=options
        )
        if not json_file_name:
            return

        sql = '''
            SELECT * FROM medicine
            ORDER BY MedicineKey
        '''
        rows = self.database.select_record(sql)

        json_data = db_utils.mysql_to_json(rows)
        text_file = open(json_file_name, "w", encoding='utf8')
        text_file.write(str(json_data))
        text_file.close()

        system_utils.show_message_box(
            QMessageBox.Information,
            'JSON資料匯出完成',
            f'<h3>{json_file_name}匯出完成.</h3>',
            'JSON 檔案格式.'
        )

    def _import_ins_exam_csv(self):
        options = QFileDialog.Options()

        options |= QFileDialog.DontUseNativeDialog
        file_name, _ = QFileDialog.getOpenFileName(
            self, "匯入健保檢驗資料",
            'exam.csv',
            "CSV檔案 (*);;csv檔 (*.csv)", options=options
        )
        if not file_name:
            return

        self._import_ins_exam(file_name)

    def _import_ins_exam(self, file_name):
        with open(file_name, encoding='utf8', newline='') as f:
            for i, l in enumerate(f):
                pass

        row_count = i + 1
        progress_dialog = QtWidgets.QProgressDialog(
            '正在轉入健保檢驗資料檔中, 請稍後...', '取消', 0, row_count, self
        )

        fields = [
            'MedicineType', 'MedicineMode', 'MedicineCode',
            'InsCode', 'MedicineName', 'MedicineAlias', 'Unit', 'Description',
        ]

        with open(file_name, encoding='utf8', newline='') as csv_file:
            rows = csv.DictReader(csv_file)

            progress_dialog.setWindowModality(QtCore.Qt.WindowModal)
            row_no = 0
            for row in rows:
                progress_dialog.setValue(row_no)
                if progress_dialog.wasCanceled():
                    break

                row_no += 1
                try:
                    self._insert_ins_exam_row(fields, row)
                except Exception:
                    pass

            progress_dialog.setValue(row_count)
            progress_dialog.deleteLater()

        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Warning)
        msg_box.setWindowTitle('檔案匯入完成')
        msg_box.setText("<font size='4'><b>健保檢驗資料檔匯入完成.</b></font>")
        msg_box.setInformativeText("請按確定鍵結束.")
        msg_box.addButton(QPushButton("確定"), QMessageBox.YesRole)
        msg_box.exec_()

    def _insert_ins_exam_row(self, fields, row):
        reference_male = row['參考值(男)']
        reference_female = row['參考值(女)']
        description = f'參考值(男): {reference_male}\n參考值(女): {reference_female}'
        medicine_mode = row['項目名稱'].replace(' ', '')

        data = [
            '檢驗',
            medicine_mode,
            row['亞東代碼'],
            row['健保碼'],
            row['報告中文名稱'],
            row['英文學名'],
            row['單位'],
            description,
        ]

        self.database.insert_record('medicine', fields, data)

    def _import_self_exam_csv(self):
        options = QFileDialog.Options()

        options |= QFileDialog.DontUseNativeDialog
        file_name, _ = QFileDialog.getOpenFileName(
            self, "匯入自費檢驗資料",
            'exam_own_expense.csv',
            "CSV檔案 (*);;csv檔 (*.csv)", options=options
        )
        if not file_name:
            return

        self._import_self_exam(file_name)

    def _import_self_exam(self, file_name):
        with open(file_name, encoding='utf8', newline='') as f:
            for i, l in enumerate(f):
                pass

        row_count = i + 1
        progress_dialog = QtWidgets.QProgressDialog(
            '正在轉入自費檢驗資料檔中, 請稍後...', '取消', 0, row_count, self
        )

        fields = [
            'MedicineType', 'MedicineMode', 'MedicineName', 'SalePrice',
        ]

        with open(file_name, encoding='utf8', newline='') as csv_file:
            rows = csv.DictReader(csv_file)

            progress_dialog.setWindowModality(QtCore.Qt.WindowModal)
            row_no = 0
            for row in rows:
                progress_dialog.setValue(row_no)
                if progress_dialog.wasCanceled():
                    break

                row_no += 1
                try:
                    self._insert_self_exam_row(fields, row)
                except Exception:
                    pass

            progress_dialog.setValue(row_count)
            progress_dialog.deleteLater()

        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Warning)
        msg_box.setWindowTitle('檔案匯入完成')
        msg_box.setText("<font size='4'><b>自費檢驗資料檔匯入完成.</b></font>")
        msg_box.setInformativeText("請按確定鍵結束.")
        msg_box.addButton(QPushButton("確定"), QMessageBox.YesRole)
        msg_box.exec_()

    def _insert_self_exam_row(self, fields, row):
        data = [
            '檢驗',
            row['類別'].strip(),
            row['項目'].replace('□', '').strip(),
            row['進價'],
        ]

        self.database.insert_record('medicine', fields, data)

    def _export_dict_medicine_excel(self):
        current_tab = self.ui.tabWidget_medicine.widget(0)
        current_tab.export_all_medicine_to_excel()

    def _import_compound_json(self):
        options = QFileDialog.Options()

        options |= QFileDialog.DontUseNativeDialog
        filename, _ = QFileDialog.getOpenFileName(
            self, "匯入成方資料",
            'compound.json',
            "JSON檔案 (*);;json檔 (*.json)", options=options
        )
        if not filename:
            return

        self._import_compound(filename)

    """
        row['方名'])
        row['出典'])
        row['效能'])
        row['適應症'])
        row['注意事項'])
        row['處方']:
            item['藥名'], item['劑量']
    """
    def _import_compound(self, filename):
        self.database.exec_sql('DELETE FROM medicine WHERE MedicineType = "成方"')
        self.database.exec_sql('DELETE FROM refcompound')

        field = ['MedicineType', 'MedicineName', 'Unit']
        compound_field = ['CompoundKey', 'MedicineKey', 'Quantity', 'Unit']

        with open(filename, 'r', encoding='utf8') as json_file:
            rows = json.load(json_file)
            for row in rows:
                compound_name = row['方名']
                data = ['成方', compound_name, '帖']
                compound_key = self.database.insert_record('medicine', field, data)

                for item in row['處方']:
                    medicine_name = item['藥名']
                    medicine_key = self._get_medicine_key('單方', medicine_name)
                    if medicine_key is None:
                        continue

                    quantity = item['劑量']
                    # quantity = round(quantity / 3.75, 2)
                    data = [compound_key, medicine_key, quantity, '克']
                    self.database.insert_record('refcompound', compound_field, data)

        self.tab_compound.read_medicine()
        system_utils.show_message_box(
            QMessageBox.Information,
            '轉檔完成',
            '<h3>成方資料轉檔完成.</h3>',
            '完成.'
        )

    def _get_medicine_key(self, medicine_type, medicine_name):
        sql = f'''
            SELECT MedicineKey FROM medicine
            WHERE
                MedicineType = "{medicine_type}" AND
                MedicineName LIKE "%{medicine_name}%"
        '''
        rows = self.database.select_record(sql)
        if len(rows) <= 0:
            return None

        row = rows[0]
        medicine_key = row['MedicineKey']

        return medicine_key

    def _import_supplier_medicine(self):
        options = QFileDialog.Options()

        options |= QFileDialog.DontUseNativeDialog
        filename, _ = QFileDialog.getOpenFileName(
            self, "匯入藥廠提供處方資料",
            '*.xlsx',
            "Excel檔案 (*);;xlsx檔 (*.xlsx)", options=options
        )
        if not filename:
            return

        self._convert_supplier_medicine(filename)

    """
        row['方名'])
        row['出典'])
        row['效能'])
        row['適應症'])
        row['注意事項'])
        row['處方']:
            item['藥名'], item['劑量']
    """
    def _convert_supplier_medicine(self, excel_filename):
        wb = openpyxl.load_workbook(excel_filename, data_only=True)
        sheet = wb[wb.sheetnames[0]]

        max_progress = sheet.max_row
        progress_dialog = QtWidgets.QProgressDialog(
            '正在匯入處方資料Excel中, 請稍後...', '取消', 0, max_progress, self
        )

        progress_dialog.setWindowModality(QtCore.Qt.WindowModal)
        progress_dialog.setValue(0)

        start_no = 2
        i = 0
        unconvertion_list = []
        for row_no in range(start_no, sheet.max_row + start_no+1):
            i += 1
            progress_dialog.setValue(i)

            medicine_name = string_utils.xstr(sheet.cell(row_no, 2).value).strip()
            if medicine_name == '':
                continue

            medicine_name = re.sub(r'\(.*?\)', '', medicine_name)
            location = string_utils.xstr(sheet.cell(row_no, 3).value).strip()

            sql = f'''
                SELECT MedicineKey FROM medicine
                WHERE
                    MedicineType IN ("單方", "複方", "自費科中") AND
                    MedicineName = "{medicine_name}"
            '''
            rows = self.database.select_record(sql)
            if len(rows) <= 0:
                unconvertion_list.append(f'{medicine_name}: {location}')
                continue

            medicine_key = rows[0]['MedicineKey']
            sql = f'''
                UPDATE medicine
                    SET Location = "{location}"
                WHERE
                    MedicineKey = {medicine_key}
            '''
            self.database.exec_sql(sql)

        wb.close()

        progress_dialog.setValue(max_progress)
        progress_dialog.deleteLater()

        system_utils.show_message_box(
            QMessageBox.Information,
            '轉檔完成',
            f'''<h3>藥廠處方資料轉檔完成.</h3>
            <h4>無法轉換的處方如下:</h4>
            <h4>{'<br>'.join(unconvertion_list)}</h4>
            ''',
            '完成.'
        )
