
# -*- coding: UTF-8 -*-

import json

import openpyxl
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import (QFileDialog, QInputDialog, QLineEdit, QMessageBox,
                             QPushButton)

from libs import (class_utils, db_utils, dialog_utils, export_utils,
                  number_utils, personnel_utils, string_utils, system_utils,
                  ui_utils)


# 處方詞庫 2019.02.25
class DictDrug(QtWidgets.QMainWindow):
    # 初始化
    def __init__(self, parent=None, *args):
        super(DictDrug, self).__init__(parent)
        self.parent = parent
        self.database = args[0]
        self.system_settings = args[1]
        self.ui = None
        self.dict_type = '藥品'

        self.user_name = system_utils.get_user_name(self.system_settings)

        self.groups_col_no = {
            'dict_groups_key': 0,
            'dict_order_no': 1,
            'dict_groups_name': 2,
            'dict_groups_percent': 3,
            'dict_groups_field_name': 4,
            'dict_groups_no_purchase': 5,
            
        }

        self.medicine_col_no = {
            'medicine_key': 0,
            'medicine_type': 1,
            'medicine_code': 2,
            'input_code': 3,
            'medicine_name': 4,
            'ins_code': 5,
            'unit': 6,
            'dosage': 7,
            'doctor_project': 8,
            'project': 9,
            'medicine_alias': 10,
            'location': 11,
            'in_price': 12,
            'sale_price': 13,
            'quantity': 14,
            'safe_quantity': 15,
            'commission': 16,
            'no_discount': 17,
            'deactivate': 18,
        }

        self._set_ui()
        self._set_signal()
        self._read_drug()
        self._set_permission()

    # 解構
    def __del__(self):
        self.close_all()

    # 關閉
    def close_all(self):
        pass

    # 設定GUI
    def _set_ui(self):
        self.ui = ui_utils.load_ui_file(ui_utils.UI_DICT_DRUG, self)
        system_utils.set_css(self, self.system_settings)
        system_utils.center_window(self)
        self.table_widget_dict_groups = class_utils.get_table_widget(self.ui.tableWidget_dict_groups, self.database)
        self.table_widget_dict_groups.set_column_hidden([0, 1])
        self.table_widget_dict_drug = class_utils.get_table_widget(self.ui.tableWidget_dict_drug, self.database)
        self.table_widget_in_price = class_utils.get_table_widget(self.ui.tableWidget_in_price, self.database)
        self.table_widget_in_price.set_column_hidden([0])
        # self.table_widget_dict_drug.set_column_hidden([0])
        self._set_table_width()
        self.ui.statusbar.showMessage(
            '若藥品類別的自費批價欄為空白, 除了「水藥」自動歸類為「水藥費」, 「高貴」歸類為「高貴藥費」, 其他均歸為「自費藥費」.'
        )

    def _set_permission(self):
        if self.user_name == '超級使用者':
            return

        if personnel_utils.get_permission(self.database, '處方資料', '更改抽成', self.user_name) != 'Y':
            self.table_widget_dict_groups.set_column_hidden([3])
            self.table_widget_dict_drug.set_column_hidden([14])

        if personnel_utils.get_permission(self.database, '系統作業', '關閉匯出功能', self.user_name) == 'Y':
            self.ui.toolButton_export.setEnabled(False)
            self.ui.toolButton_export_excel.setEnabled(False)

    # 設定信號
    def _set_signal(self):
        self.ui.tableWidget_dict_groups.itemSelectionChanged.connect(self.dict_groups_changed)
        self.ui.toolButton_add_dict_groups.clicked.connect(self._add_dict_groups)
        self.ui.toolButton_remove_dict_groups.clicked.connect(self._remove_dict_groups)
        self.ui.toolButton_edit_dict_groups.clicked.connect(self._edit_dict_groups)
        self.ui.toolButton_charge_field.clicked.connect(self._add_charge_field)
        self.ui.toolButton_edit_percent.clicked.connect(self._edit_percent)
        self.ui.tableWidget_dict_groups.doubleClicked.connect(self._edit_dict_groups)
        self.ui.toolButton_add_drug.clicked.connect(self._add_drug)
        self.ui.toolButton_remove_drug.clicked.connect(self._remove_drug)
        self.ui.toolButton_edit_drug.clicked.connect(self._edit_drug)
        self.ui.toolButton_copy_drug.clicked.connect(self._copy_drug)
        self.ui.toolButton_cut_drug.clicked.connect(self._cut_drug)
        self.ui.tableWidget_dict_drug.doubleClicked.connect(self._edit_drug)
        self.ui.tableWidget_dict_drug.itemSelectionChanged.connect(self.dict_drug_changed)
        self.ui.toolButton_search_drug.clicked.connect(self._search_drug)
        self.ui.toolButton_up.clicked.connect(self._groups_order_up)
        self.ui.toolButton_down.clicked.connect(self._groups_order_down)
        self.ui.toolButton_no_purchase.clicked.connect(self._no_purchase)
        self.ui.toolButton_export.clicked.connect(self._export_medicine)
        self.ui.toolButton_export_excel.clicked.connect(self._export_excel)
        self.ui.toolButton_import_excel.clicked.connect(self._import_excel)
        self.ui.toolButton_append_excel.clicked.connect(self._append_excel)
        self.ui.toolButton_import.clicked.connect(self._import_medicine)
        self.ui.toolButton_activate.clicked.connect(self._activate_drug)
        self.ui.toolButton_deactivate.clicked.connect(self._deactivate_drug)
        self.ui.checkBox_deactivate.clicked.connect(self._filter_deactivate_medicine)
        self.ui.checkBox_location.clicked.connect(self._filter_location_medicine)
        self.ui.toolButton_sync_in_price.clicked.connect(self._sync_in_price)
        self.ui.toolButton_sync_drug.clicked.connect(self._sync_drug)
        # self.ui.lineEdit_search_drug.textChanged.connect(self._search_drug)
        self.ui.lineEdit_search_drug.keyPressEvent = self._line_edit_search_drug_package_key_press

    def _line_edit_search_drug_package_key_press(self, event):
        key = event.key()
        if key == QtCore.Qt.Key_Return or key == QtCore.Qt.Key_Enter:
            self._search_drug()

        return QtWidgets.QLineEdit.keyPressEvent(self.ui.lineEdit_search_drug, event)

    # 設定欄位寬度
    def _set_table_width(self):
        dict_groups_width = [100, 50, 115, 60, 100, 80]
        self.table_widget_dict_groups.set_table_heading_width(dict_groups_width)

        dict_drug_width = [100, 100, 120, 70, 250, 90, 50, 50, 80, 80, 120, 80, 80, 80, 70, 70, 60, 80, 120]
        self.table_widget_dict_drug.set_table_heading_width(dict_drug_width)

        in_price_width = [100, 130, 250, 80, 90]
        self.table_widget_in_price.set_table_heading_width(in_price_width)

    def _read_drug(self):
        self._read_dict_groups()
        self._check_dict_groups_order_no()

    def _check_dict_groups_order_no(self):
        index = 0
        for row_no in range(self.ui.tableWidget_dict_groups.rowCount()):
            order_no = self.ui.tableWidget_dict_groups.item(row_no, 1)
            if order_no is not None and order_no.text() != '':
                continue

            index += 1
            dict_groups_key = self.ui.tableWidget_dict_groups.item(row_no, 0).text()
            sql = f'''
                UPDATE dict_groups
                SET
                    DictOrderNo = "{index:0>4}"
                WHERE
                    DictGroupsKey = {dict_groups_key}
            '''
            self.database.exec_sql(sql)

        self._read_dict_groups()

    def _read_dict_groups(self):
        sql = f'''
            SELECT * FROM dict_groups
            WHERE
                DictGroupsType = "{self.dict_type}類別"
            ORDER BY DictOrderNo, DictGroupsKey
        '''
        self.table_widget_dict_groups.set_db_data(sql, self._set_dict_groups_data)

    def _set_dict_groups_data(self, row_no, row):
        if row['DictGroupsLevel2'] is None:
            percent = ''
        else:
            try:
                value = number_utils.get_integer(row['DictGroupsLevel2'])
                percent = f'{value}%'
            except ValueError:
                percent = string_utils.xstr(row['DictGroupsLevel2'])

        dict_groups_row = [
            string_utils.xstr(row['DictGroupsKey']),
            string_utils.xstr(row['DictOrderNo']),
            string_utils.xstr(row['DictGroupsName']),
            percent,
            string_utils.xstr(row['DictGroupsTopLevel']),
            string_utils.xstr(row['DictGroupsLevel3']),
        ]

        for column in range(len(dict_groups_row)):
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, dict_groups_row[column])
            self.ui.tableWidget_dict_groups.setItem(
                row_no, column, item,
            )
            if column in [3]:
                self.ui.tableWidget_dict_groups.item(row_no, column).setTextAlignment(
                    QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter
                )
            elif column in [5]:
                self.ui.tableWidget_dict_groups.item(row_no, column).setTextAlignment(
                    QtCore.Qt.AlignCenter | QtCore.Qt.AlignVCenter
                )

    def dict_groups_changed(self):
        dict_groups_type = self.table_widget_dict_groups.field_value(self.groups_col_no['dict_groups_name'])
        self.ui.toolButton_remove_dict_groups.setEnabled(True)
        self.ui.toolButton_edit_dict_groups.setEnabled(True)
        if dict_groups_type in ['單方', '複方',]:
            self.ui.toolButton_remove_dict_groups.setEnabled(False)
            self.ui.toolButton_edit_dict_groups.setEnabled(False)

        self._read_dict_drug(dict_groups_type)
        self.ui.tableWidget_dict_groups.setFocus(True)
        self._set_drug_tool_buttons()

    def _set_drug_tool_buttons(self):
        if self.ui.tableWidget_dict_drug.rowCount() <= 0:
            enabled = False
        else:
            enabled = True

        self.ui.toolButton_remove_drug.setEnabled(enabled)
        self.ui.toolButton_edit_drug.setEnabled(enabled)
        self.ui.toolButton_copy_drug.setEnabled(enabled)

    def _read_dict_drug(self, dict_groups_type, keyword=None):
        if dict_groups_type in [None, '']:
            medicine_type_condition = ''
        else:
            medicine_type_condition = f'MedicineType = "{dict_groups_type}"'

        if keyword in [None, '']:
            keyword_condition = ''
        else:
            keyword_condition = f'{keyword}'

        if medicine_type_condition != '' and keyword_condition != '':
            keyword_condition = ' AND ' + keyword_condition

        # 藥品類別內可能也會有「處置」的分類, 所以藥品「處置」會被排除 2022.03.08 中和新生堂
        sql = f'''
            SELECT * FROM medicine
            WHERE
                MedicineType NOT IN ("穴道", "處置", "檢驗", "成方") AND
                {medicine_type_condition}
                {keyword_condition}
            ORDER BY FIELD(MedicineType, "單方", "複方", "水藥", "外用", "高貴", "器材"), MedicineName
        '''

        self.table_widget_dict_drug.set_db_data(sql, self._set_dict_drug_data)
        medicine_key = self.table_widget_dict_drug.field_value(self.medicine_col_no['medicine_key'])
        self._read_drug_description(medicine_key)

    def _set_dict_drug_data(self, row_no, row):
        deactivate = string_utils.xstr(row['Deactivate'])

        dict_drug_row = [
            string_utils.xstr(row['MedicineKey']),
            string_utils.xstr(row['MedicineType']),
            string_utils.xstr(row['MedicineCode']),
            string_utils.xstr(row['InputCode']),
            string_utils.xstr(row['MedicineName']),
            string_utils.xstr(row['InsCode']),
            string_utils.xstr(row['Unit']),
            number_utils.get_float(row['Dosage']),
            string_utils.xstr(row['DoctorProject']),
            string_utils.xstr(row['Project']),
            string_utils.xstr(row['MedicineAlias']),
            string_utils.xstr(row['Location']),
            number_utils.get_float(row['InPrice']),
            number_utils.get_float(row['SalePrice']),
            number_utils.get_integer(row['Quantity']),
            number_utils.get_integer(row['SafeQuantity']),
            string_utils.xstr(row['Commission']),
            string_utils.xstr(row['Charged']),
            deactivate,
        ]

        for col_no in range(len(dict_drug_row)):
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, dict_drug_row[col_no])
            self.ui.tableWidget_dict_drug.setItem(row_no, col_no, item)
            if col_no in [
                self.medicine_col_no['dosage'],
                self.medicine_col_no['in_price'],
                self.medicine_col_no['sale_price'],
                self.medicine_col_no['quantity'],
                self.medicine_col_no['safe_quantity'],
                self.medicine_col_no['commission']
            ]:
                self.ui.tableWidget_dict_drug.item(row_no, col_no).setTextAlignment(
                    QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter
                )
            elif col_no in [
                self.medicine_col_no['unit'],
                self.medicine_col_no['no_discount'],
            ]:
                self.ui.tableWidget_dict_drug.item(row_no, col_no).setTextAlignment(
                    QtCore.Qt.AlignCenter | QtCore.Qt.AlignVCenter
                )
            if deactivate != '':
                self.ui.tableWidget_dict_drug.item(row_no, col_no).setForeground(
                    QtGui.QColor('gray')
                )

    # 新增處方類別
    def _add_dict_groups(self):
        input_dialog = dialog_utils.get_dialog(
            f'{self.dict_type}類別',
            f'請輸入{self.dict_type}類別',
            None, QInputDialog.TextInput, 320, 200
        )
        if not input_dialog.exec_():
            return

        dict_groups = input_dialog.textValue()
        if dict_groups == '':
            return

        input_dialog = dialog_utils.get_dialog(
            f'{self.dict_type}類別',
            f'請輸入「{dict_groups}類別」的抽成',
            None, QInputDialog.TextInput, 320, 200
        )
        if not input_dialog.exec_():
            return

        dict_groups_percent = input_dialog.textValue().strip('%')

        sql = f'''
            SELECT * FROM dict_groups
            WHERE
                DictGroupsType = "{self.dict_type}類別"
            ORDER BY DictOrderNo DESC LIMIT 1
        '''
        rows = self.database.select_record(sql)
        last_dict_order_no = number_utils.get_integer(rows[0]['DictOrderNo'])

        field = [
            'DictGroupsType', 'DictOrderNo', 'DictGroupsName', 'DictGroupsLevel2'
        ]

        data = [
            f'{self.dict_type}類別',
            last_dict_order_no+1,
            dict_groups,
            dict_groups_percent,
        ]
        self.database.insert_record('dict_groups', field, data)
        self._read_dict_groups()

    # 移除處方類別
    def _remove_dict_groups(self):
        dict_groups = self.table_widget_dict_groups.field_value(self.groups_col_no['dict_groups_name'])
        msg_box = dialog_utils.get_message_box(
            f'刪除{self.dict_type}類別資料',
            QMessageBox.Warning,
            f'<font size="5" color="red"><b>確定刪除 [{dict_groups}] {self.dict_type}類別?</b></font>',
            '注意！資料刪除後, 將無法回復!'
        )
        remove_record = msg_box.exec_()
        if not remove_record:
            return

        key = self.table_widget_dict_groups.field_value(self.groups_col_no['dict_groups_key'])
        self.database.delete_record('dict_groups', 'DictGroupsKey', key)
        self.ui.tableWidget_dict_groups.removeRow(self.ui.tableWidget_dict_groups.currentRow())

    # 更改處方類別抽成
    def _edit_percent(self):
        dict_groups_key = self.table_widget_dict_groups.field_value(self.groups_col_no['dict_groups_key'])
        if dict_groups_key == '':
            return

        dict_groups = self.table_widget_dict_groups.field_value(self.groups_col_no['dict_groups_name'])
        old_percent = self.table_widget_dict_groups.field_value(self.groups_col_no['dict_groups_percent'])
        old_percent = old_percent.replace('%', '')
        input_dialog = dialog_utils.get_dialog(
            f'{self.dict_type}類別',
            f'請輸入「{dict_groups}類別」的抽成',
            old_percent, QInputDialog.TextInput, 320, 200
        )
        if not input_dialog.exec_():
            return

        dict_groups_percent = input_dialog.textValue().strip('%')

        field = [
            'DictGroupsType', 'DictOrderNo', 'DictGroupsName', 'DictGroupsLevel2'
        ]

        sql = f'''
            UPDATE dict_groups
            SET
                DictGroupsLevel2 = "{dict_groups_percent}"
            WHERE
                DictGroupsKey = "{dict_groups_key}"
        '''
        self.database.exec_sql(sql)
        self._read_dict_groups()

    # 更改處方類別
    def _edit_dict_groups(self):
        old_groups = self.table_widget_dict_groups.field_value(self.groups_col_no['dict_groups_name'])
        if old_groups in ['單方', '複方', '水藥', '外用', '高貴', '器材']:
            return

        input_dialog = dialog_utils.get_dialog(
            f'{self.dict_type}類別',
            f'請輸入{self.dict_type}類別',
            old_groups,
            QInputDialog.TextInput, 320, 200
        )
        if not input_dialog.exec_():
            return

        dict_groups_name = input_dialog.textValue()
        if dict_groups_name == '':
            return

        old_percent = self.table_widget_dict_groups.field_value(
            self.groups_col_no['dict_groups_percent']).strip('%')

        input_dialog = dialog_utils.get_dialog(
            f'{self.dict_type}類別',
            f'請輸入「{dict_groups_name}類別」的抽成',
            old_percent,
            QInputDialog.TextInput, 320, 200
        )
        if not input_dialog.exec_():
            return

        dict_groups_percent = input_dialog.textValue().strip('%')

        data = [
            dict_groups_name,
            dict_groups_percent,
        ]

        sql = f'''
            UPDATE dict_groups
            SET
                DictGroupsTopLevel = "{dict_groups_name}"
            WHERE
                DictGroupsType = "{self.dict_type}" AND
                DictGroupsTopLevel = "{old_groups}"
        '''
        self.database.exec_sql(sql)

        fields = ['DictGroupsName', 'DictGroupsLevel2']
        self.database.update_record(
            'dict_groups', fields, 'DictGroupsKey',
            self.table_widget_dict_groups.field_value(self.groups_col_no['dict_groups_key']), data
        )

        dict_groups_key = self.table_widget_dict_groups.field_value(self.groups_col_no['dict_groups_key'])
        sql = f'''
            SELECT * FROM dict_groups
            WHERE
                DictGroupsKey = {dict_groups_key}
        '''
        row = self.database.select_record(sql)[0]
        self._set_dict_groups_data(self.ui.tableWidget_dict_groups.currentRow(), row)

        sql = f'''
            UPDATE medicine
            SET
                MedicineType = "{dict_groups_name}"
            WHERE
                MedicineType = "{old_groups}"
        '''
        self.database.exec_sql(sql)

    # 更改處方類別
    def _add_charge_field(self):
        items = ['自費藥費', '水藥費', '高貴藥費', '自費針灸費', '民俗調理費', '自費材料費', '自費診察費']

        dict_groups_key = self.table_widget_dict_groups.field_value(self.groups_col_no['dict_groups_key'])
        dict_groups = self.table_widget_dict_groups.field_value(self.groups_col_no['dict_groups_name'])
        charge_field = self.table_widget_dict_groups.field_value(self.groups_col_no['dict_groups_field_name'])

        try:
            index = items.index(charge_field)
        except ValueError:
            index = 0

        item, ok = QInputDialog.getItem(
            self,
            f'{self.dict_type}類別',
            f'請選擇「{dict_groups}類別」的自費批價欄位',
            items, index, False
        )

        if not ok:
            return

        sql = f'''
            UPDATE dict_groups
            SET
                DictGroupsTopLevel = "{item}"
            WHERE
                DictGroupsKey = {dict_groups_key}
        '''
        self.database.exec_sql(sql)
        self._read_dict_groups()

    # 新增處方
    def _add_drug(self):
        current_row = self.ui.tableWidget_dict_drug.rowCount()
        self.ui.tableWidget_dict_drug.insertRow(current_row)
        dict_groups_type = self.table_widget_dict_groups.field_value(self.groups_col_no['dict_groups_name'])

        dialog = dialog_utils.get_dialog_input_drug(
            self, self.database, self.system_settings, dict_groups_type, None)
        dialog.exec_()
        dialog.close_all()
        dialog.deleteLater()

        self._read_dict_drug(dict_groups_type)

    def _save_drug(self, dialog):
        current_row = self.ui.tableWidget_dict_drug.rowCount()
        self.ui.tableWidget_dict_drug.insertRow(current_row)
        dict_groups_type = self.table_widget_dict_groups.field_value(self.groups_col_no['dict_groups_name'])

        if dialog.ui.checkBox_no_discount.isChecked():
            no_discount = 'Y'
        else:
            no_discount = None

        fields = [
            'MedicineType', 'MedicineCode', 'InputCode', 'MedicineName', 'Unit', 'MedicineMode', 'InsCode',
            'Dosage', 'MedicineAlias', 'Location', 'InPrice', 'SalePrice', 'Commission',
            'Quantity', 'SafeQuantity', 'Charged',
            'Description',
        ]
        data = [
            dict_groups_type,
            dialog.ui.lineEdit_medicine_code.text(),
            dialog.ui.lineEdit_input_code.text(),
            dialog.ui.lineEdit_medicine_name.text(),
            dialog.ui.comboBox_unit.currentText(),
            dialog.ui.comboBox_medicine_mode.currentText(),
            dialog.ui.lineEdit_ins_code.text(),
            dialog.ui.lineEdit_dosage.text(),
            dialog.ui.lineEdit_medicine_alias.text(),
            dialog.ui.lineEdit_location.text(),
            dialog.ui.lineEdit_in_price.text(),
            dialog.ui.lineEdit_sale_price.text(),
            dialog.ui.lineEdit_commission.text(),
            dialog.ui.lineEdit_quantity.text(),
            dialog.ui.lineEdit_safe_quantity.text(),
            no_discount,
            dialog.ui.textEdit_description.toPlainText(),
        ]
        string_utils.str_to_none(data)
        medicine_key = self.database.insert_record('medicine', fields, data)
        self._read_dict_drug(dict_groups_type)

        return medicine_key

    def _save_commission(self, dialog, medicine_key):
        fields = ['MedicineKey', 'Name', 'Commission', 'Remark']

        for row_no in range(dialog.ui.tableWidget_commission.rowCount()):
            name = dialog.ui.tableWidget_commission.item(row_no, 0).text()
            commission = dialog.ui.tableWidget_commission.item(row_no, 1).text()
            remark = dialog.ui.tableWidget_commission.item(row_no, 2).text()

            data = [medicine_key, name, commission, remark]
            self.database.insert_record('commission', fields, data)

    # 移除處方
    def _remove_drug(self):
        selected = self.ui.tableWidget_dict_drug.selectedRanges()
        drug_list = []
        for item in selected:
            for row_no in range(item.topRow(), item.bottomRow() + 1):
                drug_list.append(self.ui.tableWidget_dict_drug.item(row_no, 0).text())

        if len(drug_list) > 1:
            self._remove_multiple_drugs(drug_list)
        else:
            self._remove_single_drug()

    def _remove_multiple_drugs(self, drug_list):
        msg_box = dialog_utils.get_message_box(
            f'刪除{self.dict_type}資料',
            QMessageBox.Warning,
            '<font size="5" color="red"><b>確定刪除選取的處方資料?</b></font>',
            '注意！資料刪除後, 將無法回復!'
        )

        remove_record = msg_box.exec_()
        if not remove_record:
            return

        drug_list = str(drug_list)[1:-1]
        sql = f'''
            DELETE FROM medicine
            WHERE
                MedicineKey IN ({drug_list})
        '''
        self.database.exec_sql(sql)
        dict_groups_type = self.table_widget_dict_groups.field_value(self.groups_col_no['dict_groups_name'])
        self._read_dict_drug(dict_groups_type)

    def _remove_single_drug(self):
        drug_name = self.table_widget_dict_drug.field_value(self.medicine_col_no['medicine_name'])
        msg_box = dialog_utils.get_message_box(
            f'刪除{self.dict_type}資料',
            QMessageBox.Warning,
            f'<font size="5" color="red"><b>確定刪除{self.dict_type}: {drug_name}?</b></font>',
            '注意！資料刪除後, 將無法回復!'
        )
        remove_record = msg_box.exec_()
        if not remove_record:
            return

        key = self.table_widget_dict_drug.field_value(self.medicine_col_no['medicine_key'])
        self.database.delete_record('medicine', 'MedicineKey', key)
        self.database.delete_record('commission', 'MedicineKey', key)
        self.ui.tableWidget_dict_drug.removeRow(self.ui.tableWidget_dict_drug.currentRow())
        self._set_drug_tool_buttons()

    # 更改處方
    def _edit_drug(self):
        medicine_key = self.table_widget_dict_drug.field_value(self.medicine_col_no['medicine_key'])
        dialog = dialog_utils.get_dialog_input_drug(
            self, self.database, self.system_settings, None, medicine_key)
        dialog.exec_()
        dialog.close_all()
        dialog.deleteLater()

        sql = f'''
            SELECT * FROM medicine
            WHERE
                MedicineKey = {medicine_key}
        '''
        row_data = self.database.select_record(sql)[0]
        self._set_dict_drug_data(self.ui.tableWidget_dict_drug.currentRow(), row_data)

    # 複製處方
    def _copy_drug(self):
        dict_groups_type = self.table_widget_dict_groups.field_value(self.groups_col_no['dict_groups_name'])

        sql = '''
            SELECT * FROM dict_groups
            WHERE
                DictGroupsType IN ("藥品類別", "處置類別")
            ORDER BY DictGroupsType, DictOrderNo
        '''
        rows = self.database.select_record(sql)
        items = ()
        for row in rows:
            dict_groups_name = string_utils.xstr(row['DictGroupsName'])
            if dict_groups_name == dict_groups_type:
                continue

            items += (dict_groups_name, )

        medicine_type, ok = QInputDialog.getItem(
            self, "拷貝處方詞庫", "請選擇拷貝到何處", items, 0, False
        )

        if not ok:
            return

        selected = self.ui.tableWidget_dict_drug.selectedRanges()
        drug_list = []
        for item in selected:
            for row_no in range(item.topRow(), item.bottomRow() + 1):
                drug_list.append(self.ui.tableWidget_dict_drug.item(row_no, 0).text())

        fields = [
            'MedicineType', 'MedicineMode', 'MedicineCode', 'InputCode', 'InsCode',
            'MedicineName', 'MedicineAlias', 'Unit', 'Dosage', 'Location',
            'SalePrice', 'InPrice', 'Charged', 'SafeQuantity', 'Description',
        ]
        for medicine_key in drug_list:
            sql = f'''
                SELECT * FROM medicine
                WHERE
                    MedicineKey = {medicine_key}
            '''
            rows = self.database.select_record(sql)
            if len(rows) <= 0:
                continue

            row = rows[0]
            data = [
                medicine_type,
                string_utils.xstr(row['MedicineMode']),
                string_utils.xstr(row['MedicineCode']),
                string_utils.xstr(row['InputCode']),
                string_utils.xstr(row['InsCode']),
                string_utils.xstr(row['MedicineName']),
                string_utils.xstr(row['MedicineAlias']),
                string_utils.xstr(row['Unit']),
                string_utils.xstr(row['Dosage']),
                string_utils.xstr(row['Location']),
                string_utils.xstr(row['SalePrice']),
                string_utils.xstr(row['InPrice']),
                string_utils.xstr(row['Charged']),
                string_utils.xstr(row['SafeQuantity']),
                string_utils.get_str(row['Description'], 'utf8'),
            ]
            self.database.insert_record('medicine', fields, data)

        if dict_groups_type == medicine_type:
            self._read_dict_drug(dict_groups_type)

    # 剪下處方
    def _cut_drug(self):
        dict_groups_type = self.table_widget_dict_groups.field_value(self.groups_col_no['dict_groups_name'])
        sql = '''
            SELECT * FROM dict_groups
            WHERE
                DictGroupsType IN ("藥品類別", "處置類別")
            ORDER BY DictGroupsType, DictOrderNo
        '''

        rows = self.database.select_record(sql)
        items = ()
        for row in rows:
            dict_groups_name = string_utils.xstr(row['DictGroupsName'])
            if dict_groups_name == dict_groups_type:
                continue

            items += (dict_groups_name, )

        medicine_type, ok = QInputDialog.getItem(
            self, "剪下處方詞庫", "請選擇貼到到何處", items, 0, False
        )

        if not ok:
            return

        selected = self.ui.tableWidget_dict_drug.selectedRanges()
        drug_list = []
        for item in selected:
            for row_no in range(item.topRow(), item.bottomRow() + 1):
                drug_list.append(self.ui.tableWidget_dict_drug.item(row_no, 0).text())

        for medicine_key in drug_list:
            sql = f'''
                UPDATE medicine
                SET
                    MedicineType = "{medicine_type}"
                WHERE
                    MedicineKey = {medicine_key}
            '''
            self.database.exec_sql(sql)

        self._read_dict_drug(dict_groups_type)

    def dict_drug_changed(self):
        medicine_key = self.table_widget_dict_drug.field_value(self.medicine_col_no['medicine_key'])
        self._read_drug_description(medicine_key)
        self._read_in_price(medicine_key)
        self.ui.tableWidget_dict_drug.setFocus()

        self._set_drug_tool_buttons()

    def _read_drug_description(self, medicine_key):
        self.ui.textEdit_description.setText('')

        sql = f'''
            SELECT * FROM medicine
            WHERE
                MedicineKey = "{medicine_key}"
        '''
        rows = self.database.select_record(sql)
        if len(rows) <= 0:
            return

        try:
            self.ui.textEdit_description.setText(string_utils.get_str(rows[0]['Description'], 'utf8'))
        except TypeError:
            pass

    def _read_in_price(self, medicine_key):
        sql = f'''
            SELECT stockinitems.*, stockin.* FROM stockinitems
                LEFT JOIN stockin ON stockinitems.StockInKey = stockin.StockInKey
            WHERE
                MedicineKey = "{medicine_key}"
            ORDER BY stockin.StockInDate DESC
        '''
        self.table_widget_in_price.set_db_data(sql, self._set_price_in_data)

    def _set_price_in_data(self, row_no, row):
        price_in_row = [
            string_utils.xstr(row['StockInItemsKey']),
            string_utils.xstr(row['StockInDate']),
            string_utils.xstr(row['Supplier']),
            string_utils.xstr(row['Unit']),
            number_utils.get_float(row['UnitPrice']),
        ]

        for col_no in range(len(price_in_row)):
            item = QtWidgets.QTableWidgetItem()
            item.setData(QtCore.Qt.EditRole, price_in_row[col_no])
            self.ui.tableWidget_in_price.setItem(row_no, col_no, item)
            if col_no in [4]:
                self.ui.tableWidget_in_price.item(row_no, col_no).setTextAlignment(
                    QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter
                )
            elif col_no in [3]:
                self.ui.tableWidget_in_price.item(row_no, col_no).setTextAlignment(
                    QtCore.Qt.AlignCenter | QtCore.Qt.AlignVCenter
                )

    # 主程式控制關閉此分頁
    def close_tab(self):
        current_tab = self.parent.ui.tabWidget_window.currentIndex()
        self.parent.close_tab(current_tab)

    # 關閉分頁
    def close_charge_settings(self):
        self.close_all()
        self.close_tab()

    def _search_drug(self):
        if self.ui.checkBox_deactivate.isChecked():
            self.ui.checkBox_deactivate.setChecked(False)

        dict_groups_type = self.table_widget_dict_groups.field_value(self.groups_col_no['dict_groups_name'])
        keyword = self.ui.lineEdit_search_drug.text()

        if keyword == '':
            self._read_dict_drug(dict_groups_type)
        else:
            if self.ui.checkBox_find_all.isChecked():
                dict_groups_type = None

            script = f'''
                (InputCode LIKE "{keyword}%" OR
                 MedicineCode LIKE "{keyword}%" OR
                 MedicineName LIKE "%{keyword}%")
            '''
            self._read_dict_drug(dict_groups_type, script)

        self.ui.lineEdit_search_drug.setFocus(True)
        self.ui.lineEdit_search_drug.setCursorPosition(len(keyword))

    def _groups_order_up(self):
        current_row = self.ui.tableWidget_dict_groups.currentRow()
        if current_row == 0:
            return

        prior_dict_row = [
            self.ui.tableWidget_dict_groups.item(current_row-1, 0).text(),
            self.ui.tableWidget_dict_groups.item(current_row, 1).text(),
            self.ui.tableWidget_dict_groups.item(current_row-1, 2).text(),
        ]

        current_dict_row = [
            self.ui.tableWidget_dict_groups.item(current_row, 0).text(),
            self.ui.tableWidget_dict_groups.item(current_row-1, 1).text(),
            self.ui.tableWidget_dict_groups.item(current_row, 2).text(),
        ]

        self._update_dict_groups(prior_dict_row)
        self._update_dict_groups(current_dict_row)
        self._refresh_dict_groups_row(current_row-1, current_dict_row)
        self._refresh_dict_groups_row(current_row, prior_dict_row)

        self.ui.tableWidget_dict_groups.setCurrentCell(current_row-1, 1)

    def _groups_order_down(self):
        current_row = self.ui.tableWidget_dict_groups.currentRow()
        if current_row >= self.ui.tableWidget_dict_groups.rowCount() - 1:
            return

        next_dict_row = [
            self.ui.tableWidget_dict_groups.item(current_row+1, 0).text(),
            self.ui.tableWidget_dict_groups.item(current_row, 1).text(),
            self.ui.tableWidget_dict_groups.item(current_row+1, 2).text(),
        ]

        current_dict_row = [
            self.ui.tableWidget_dict_groups.item(current_row, 0).text(),
            self.ui.tableWidget_dict_groups.item(current_row+1, 1).text(),
            self.ui.tableWidget_dict_groups.item(current_row, 2).text(),
        ]

        self._update_dict_groups(next_dict_row)
        self._update_dict_groups(current_dict_row)
        self._refresh_dict_groups_row(current_row+1, current_dict_row)
        self._refresh_dict_groups_row(current_row, next_dict_row)

        self.ui.tableWidget_dict_groups.setCurrentCell(current_row+1, 1)

    def _update_dict_groups(self, dict_groups_row):
        dict_order_no = dict_groups_row[1]
        dict_groups_key = dict_groups_row[0]
        self.database.exec_sql(f'''
            UPDATE dict_groups
            SET
                DictOrderNo = "{dict_order_no}"
            WHERE
                DictGroupsKey = {dict_groups_key}
        ''')

    def _refresh_dict_groups_row(self, row_no, dict_groups_row):
        for col_no in range(len(dict_groups_row)):
            self.ui.tableWidget_dict_groups.setItem(
                row_no, col_no, QtWidgets.QTableWidgetItem(dict_groups_row[col_no]),
            )

    def _export_medicine(self):
        dict_groups_type = self.table_widget_dict_groups.field_value(self.groups_col_no['dict_groups_name'])

        options = QFileDialog.Options()
        json_file_name, _ = QFileDialog.getSaveFileName(
            self.parent,
            "匯出處方JSON檔案",
            f'{dict_groups_type}藥品資料.json',
            "json檔案 (*.json)",
            options=options
        )
        if not json_file_name:
            return

        selected = self.ui.tableWidget_dict_drug.selectedRanges()
        drug_list = []
        for item in selected:
            for row_no in range(item.topRow(), item.bottomRow() + 1):
                drug_list.append(self.ui.tableWidget_dict_drug.item(row_no, 0).text())

        drug_list = str(drug_list)[1:-1]
        sql = f'''
            SELECT * FROM medicine
            WHERE
                MedicineKey IN ({drug_list})
        '''
        rows = self.database.select_record(sql)

        json_data = db_utils.mysql_to_json(rows)
        text_file = open(json_file_name, "w", encoding='utf8')
        text_file.write(str(json_data))
        text_file.close()

        system_utils.show_message_box(
            QMessageBox.Information,
            'JSON資料匯出完成',
            f'<h3>處方匯出資料{json_file_name}匯出完成.</h3>',
            'JSON 檔案格式.'
        )

    def _import_medicine(self):
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getOpenFileName(
            self, "開啟處方JSON檔",
            '*.json', "json 檔 (*.json);;", options=options
        )
        if not file_name:
            return

        dict_groups_type = self.table_widget_dict_groups.field_value(self.groups_col_no['dict_groups_name'])
        fields = [
            'MedicineType', 'MedicineCode', 'InputCode', 'MedicineName', 'Unit', 'MedicineMode', 'InsCode',
            'Dosage', 'MedicineAlias', 'Location', 'InPrice', 'SalePrice', 'Commission',
            'Quantity', 'SafeQuantity',
            'Description',
        ]

        with open(file_name, encoding='utf8') as json_file:
            rows = json.load(json_file)
            for row in rows:
                data = [
                    dict_groups_type,
                    row['MedicineCode'],
                    row['InputCode'],
                    row['MedicineName'],
                    row['Unit'],
                    row['MedicineMode'],
                    row['InsCode'],
                    row['Dosage'],
                    row['MedicineAlias'],
                    row['Location'],
                    row['InPrice'],
                    row['SalePrice'],
                    row['Commission'],
                    row['Quantity'],
                    row['SafeQuantity'],
                    row['Description'],
                ]
                self.database.insert_record('medicine', fields, data)

        self._read_dict_drug(dict_groups_type)

    def _filter_deactivate_medicine(self):
        for row_no in range(self.ui.tableWidget_dict_drug.rowCount(), -1, -1):
            deactivate_item = self.ui.tableWidget_dict_drug.item(row_no, self.medicine_col_no['deactivate'])
            self.ui.tableWidget_dict_drug.showRow(row_no)
            if self.ui.checkBox_deactivate.isChecked():
                if deactivate_item is None or deactivate_item.text() == '':
                    self.ui.tableWidget_dict_drug.hideRow(row_no)

    def _filter_location_medicine(self):
        for row_no in range(self.ui.tableWidget_dict_drug.rowCount(), -1, -1):
            location_item = self.ui.tableWidget_dict_drug.item(row_no, self.medicine_col_no['location'])
            self.ui.tableWidget_dict_drug.showRow(row_no)
            if self.ui.checkBox_location.isChecked():
                if location_item is None or location_item.text() == '':
                    self.ui.tableWidget_dict_drug.hideRow(row_no)

    def _export_excel(self):
        dict_groups_type = self.table_widget_dict_groups.field_value(self.groups_col_no['dict_groups_name'])

        options = QFileDialog.Options()
        excel_file_name, _ = QFileDialog.getSaveFileName(
            self.parent,
            "匯出處方Excel檔案", f'{dict_groups_type}藥品資料.xlsx',
            "excel檔案 (*.xlsx)",
            options=options
        )
        if not excel_file_name:
            return

        export_utils.export_table_widget_to_excel(
            excel_file_name, self.ui.tableWidget_dict_drug, None, [12, 13, 14, 15, 16]
        )
        system_utils.show_message_box(
            QMessageBox.Information,
            'Excel資料匯出完成',
            f'<h3>處方匯出資料{excel_file_name}匯出完成.</h3>',
            'Excel檔案格式.'
        )

    # 啟用處方
    def _activate_drug(self):
        msg_box = dialog_utils.get_message_box(
            '啟用藥品', QMessageBox.Warning,
            '<font size="5" color="red"><b>確定啟用所選取的藥品?</b></font>',
            '注意！藥品啟用後, 停用原因將設為空白!'
        )
        activate = msg_box.exec_()
        if not activate:
            return

        selected = self.ui.tableWidget_dict_drug.selectedRanges()
        for item in selected:
            for row_no in range(item.topRow(), item.bottomRow() + 1):
                medicine_key = self.ui.tableWidget_dict_drug.item(row_no, 0).text()
                sql = f'''
                    UPDATE medicine
                    SET
                        Deactivate = NULL
                    WHERE
                        MedicineKey = {medicine_key}
                '''
                self.database.exec_sql(sql)

                sql = f'''
                    SELECT * FROM medicine
                    WHERE
                        MedicineKey = {medicine_key}
                '''
                row_data = self.database.select_record(sql)[0]
                self._set_dict_drug_data(row_no, row_data)

    # 停用處方
    def _deactivate_drug(self):
        deactivate, ok = QInputDialog.getText(
            self, '停用藥品', '請輸入停用原因', QLineEdit.Normal, '缺貨')
        if not ok or deactivate == '':
            return

        selected = self.ui.tableWidget_dict_drug.selectedRanges()
        for item in selected:
            for row_no in range(item.topRow(), item.bottomRow() + 1):
                medicine_key = self.ui.tableWidget_dict_drug.item(row_no, 0).text()
                sql = f'''
                    UPDATE medicine
                    SET
                        Deactivate = "{deactivate}"
                    WHERE
                        MedicineKey = {medicine_key}
                '''
                self.database.exec_sql(sql)

                sql = f'''
                    SELECT * FROM medicine
                    WHERE
                        MedicineKey = {medicine_key}
                '''
                row_data = self.database.select_record(sql)[0]
                self._set_dict_drug_data(row_no, row_data)

    def export_all_medicine_to_excel(self):
        options = QFileDialog.Options()
        excel_file_name, _ = QFileDialog.getSaveFileName(
            self.parent,
            "匯出藥品資料Excel檔案", 'medicine.xlsx',
            "Excel檔案 (*.xlsx)",
            options=options
        )
        if not excel_file_name:
            return

        export_utils.export_multiple_table_widgets_to_excel(
            excel_file_name, self.ui.tableWidget_dict_groups, self.ui.tableWidget_dict_drug,
            None, [7, 12, 13, 14, 15, 16],
            column_width=[10, 10, 10, 10, 30, 10, 10, 10, 10, 10, 30, 15]
        )
        system_utils.show_message_box(
            QMessageBox.Information,
            'Excel資料匯出完成',
            f'<h3>處方匯出資料{excel_file_name}匯出完成.</h3>',
            'Excel檔案格式.'
        )

    def _import_excel(self):
        options = QFileDialog.Options()
        excel_filename, _ = QFileDialog.getOpenFileName(
            self, "開啟藥品檔",
            '*.xlsx', "xlsx 檔 (*.xlsx)", options=options
        )
        if not excel_filename:
            return

        wb = openpyxl.load_workbook(excel_filename, data_only=True)
        sheet = wb['sheet1']

        max_progress = sheet.max_row
        progress_dialog = QtWidgets.QProgressDialog(
            '正在匯入處方資料Excel中, 請稍後...', '取消', 0, max_progress, self
        )

        progress_dialog.setWindowModality(QtCore.Qt.WindowModal)
        progress_dialog.setValue(0)

        start_no = 2
        i = 0
        for row_no in range(start_no, sheet.max_row + start_no+1):
            i += 1
            progress_dialog.setValue(i)

            medicine_key = string_utils.xstr(sheet.cell(row_no, 1).value)
            medicine_type = string_utils.xstr(sheet.cell(row_no, 2).value)
            medicine_code = string_utils.xstr(sheet.cell(row_no, 3).value)
            input_code = string_utils.xstr(sheet.cell(row_no, 4).value)
            medicine_name = string_utils.xstr(sheet.cell(row_no, 5).value)
            ins_code = string_utils.xstr(sheet.cell(row_no, 6).value)
            unit = string_utils.xstr(sheet.cell(row_no, 7).value)
            dosage = string_utils.xstr(sheet.cell(row_no, 8).value)
            doctor_project = string_utils.xstr(sheet.cell(row_no, 9).value)
            project = string_utils.xstr(sheet.cell(row_no, 10).value)
            medicine_alias = string_utils.xstr(sheet.cell(row_no, 11).value)
            location = string_utils.xstr(sheet.cell(row_no, 12).value)
            in_price = string_utils.xstr(sheet.cell(row_no, 13).value)
            sale_price = string_utils.xstr(sheet.cell(row_no, 14).value)
            quantity = string_utils.xstr(sheet.cell(row_no, 15).value)
            safe_quantity = string_utils.xstr(sheet.cell(row_no, 16).value)
            commission = string_utils.xstr(sheet.cell(row_no, 17).value)
            no_discount = string_utils.xstr(sheet.cell(row_no, 18).value)
            deactivate = string_utils.xstr(sheet.cell(row_no, 19).value)

            row = {
                'MedicineKey': medicine_key,
                'MedicineType': medicine_type,
                'MedicineCode': medicine_code,
                'InputCode': input_code,
                'MedicineName': medicine_name,
                'InsCode': ins_code,
                'Unit': unit,
                'Dosage': dosage,
                'DoctorProject': doctor_project,
                'Project': project,
                'MedicineAlias': medicine_alias,
                'Location': location,
                'InPrice': in_price,
                'SalePrice': sale_price,
                'Quantity': quantity,
                'SafeQuantity': safe_quantity,
                'Commission': commission,
                'Charged': no_discount,
                'Deactivate': deactivate,
            }

            try:
                if medicine_key in ['', None]:
                    self._insert_medicine(row)
                else:
                    self._update_medicine(row)
            except Exception:
                print(medicine_type, medicine_name)

        wb.close()

        progress_dialog.setValue(max_progress)
        progress_dialog.deleteLater()

        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Information)
        msg_box.setWindowTitle('藥品檔匯入完成')
        msg_box.setText(f'''
            <b>藥品檔匯入完成.<br>
        ''')
        msg_box.setInformativeText("所有資料均已匯入.")
        msg_box.addButton(QPushButton("確定"), QMessageBox.YesRole)
        msg_box.exec_()

    def _append_excel(self):
        options = QFileDialog.Options()
        excel_filename, _ = QFileDialog.getOpenFileName(
            self, "開啟藥品檔",
            '*.xlsx', "xlsx 檔 (*.xlsx)", options=options
        )
        if not excel_filename:
            return

        wb = openpyxl.load_workbook(excel_filename, data_only=True)
        sheet = wb['sheet1']

        max_progress = sheet.max_row
        progress_dialog = QtWidgets.QProgressDialog(
            '正在匯入處方資料Excel中, 請稍後...', '取消', 0, max_progress, self
        )

        progress_dialog.setWindowModality(QtCore.Qt.WindowModal)
        progress_dialog.setValue(0)


        dict_medicine_type = self.table_widget_dict_groups.field_value(self.groups_col_no['dict_groups_name'])
        start_no = 2
        i = 0
        for row_no in range(start_no, sheet.max_row + start_no+1):
            i += 1
            progress_dialog.setValue(i)

            medicine_key = string_utils.xstr(sheet.cell(row_no, 1).value)
            medicine_type = string_utils.xstr(sheet.cell(row_no, 2).value)
            medicine_code = string_utils.xstr(sheet.cell(row_no, 3).value)
            input_code = string_utils.xstr(sheet.cell(row_no, 4).value)
            medicine_name = string_utils.xstr(sheet.cell(row_no, 5).value)
            ins_code = string_utils.xstr(sheet.cell(row_no, 6).value)
            unit = string_utils.xstr(sheet.cell(row_no, 7).value)
            dosage = string_utils.xstr(sheet.cell(row_no, 8).value)
            doctor_project = string_utils.xstr(sheet.cell(row_no, 9).value)
            project = string_utils.xstr(sheet.cell(row_no, 10).value)
            medicine_alias = string_utils.xstr(sheet.cell(row_no, 11).value)
            location = string_utils.xstr(sheet.cell(row_no, 12).value)
            in_price = string_utils.xstr(sheet.cell(row_no, 13).value)
            sale_price = string_utils.xstr(sheet.cell(row_no, 14).value)
            quantity = string_utils.xstr(sheet.cell(row_no, 15).value)
            safe_quantity = string_utils.xstr(sheet.cell(row_no, 16).value)
            commission = string_utils.xstr(sheet.cell(row_no, 17).value)
            no_discount = string_utils.xstr(sheet.cell(row_no, 18).value)
            deactivate = string_utils.xstr(sheet.cell(row_no, 19).value)

            row = {
                'MedicineKey': medicine_key,
                'MedicineType': medicine_type,
                'MedicineCode': medicine_code,
                'InputCode': input_code,
                'MedicineName': medicine_name,
                'InsCode': ins_code,
                'Unit': unit,
                'Dosage': dosage,
                'DoctorProject': doctor_project,
                'Project': project,
                'MedicineAlias': medicine_alias,
                'Location': location,
                'InPrice': in_price,
                'SalePrice': sale_price,
                'Quantity': quantity,
                'SafeQuantity': safe_quantity,
                'Commission': commission,
                'Charged': no_discount,
                'Deactivate': deactivate,
            }

            try:
                self._insert_medicine(row, dict_medicine_type)
            except Exception:
                print(medicine_type, medicine_name)

        wb.close()

        progress_dialog.setValue(max_progress)
        progress_dialog.deleteLater()

        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Information)
        msg_box.setWindowTitle('藥品檔匯入完成')
        msg_box.setText(f'''
            <b>藥品檔匯入完成.<br>
        ''')
        msg_box.setInformativeText("所有資料均已匯入.")
        msg_box.addButton(QPushButton("確定"), QMessageBox.YesRole)
        msg_box.exec_()
        
    def _insert_medicine(self, row, medicine_type=None):
        if medicine_type is None:
            medicine_type = row['MedicineType']
            
        fields = [
            'MedicineType', 'MedicineCode', 'InputCode', 'MedicineName', 'InsCode', 'Unit',
            'Dosage', 'DoctorProject', 'Project', 'MedicineAlias', 'Location',
            'InPrice', 'SalePrice', 'Quantity', 'SafeQuantity', 'Commission', 'Charged', 'Deactivate',
        ]
        data = [
            medicine_type, row['MedicineCode'], row['InputCode'], row['MedicineName'],
            row['InsCode'], row['Unit'], row['Dosage'], row['DoctorProject'], row['Project'],
            row['MedicineAlias'], row['Location'], row['InPrice'], row['SalePrice'],
            row['Quantity'], row['SafeQuantity'], row['Commission'], row['Charged'], row['Deactivate'],
        ]
        self.database.insert_record('medicine', fields, data)

    def _update_medicine(self, row):
        sql = f'''
            UPDATE medicine
            SET
                MedicineCode = "{row['MedicineCode']}",
                MedicineName = "{row['MedicineName']}",
                InputCode = "{row['InputCode']}",
                InsCode = "{row['InsCode']}",
                Unit = "{row['Unit']}",
                Dosage = {row['Dosage']},
                DoctorProject = "{row['DoctorProject']}",
                Project = "{row['Project']}",
                MedicineAlias = "{row['MedicineAlias']}",
                Location = "{row['Location']}",
                InPrice = {row['InPrice']},
                SalePrice = {row['SalePrice']},
                Quantity = {row['Quantity']},
                SafeQuantity = {row['SafeQuantity']},
                Commission = "{row['Commission']}",
                Charged = "{row['Charged']}",
                Deactivate = "{row['Deactivate']}"
            WHERE
                MedicineKey = {row['MedicineKey']}
        '''
        self.database.exec_sql(sql)

    def _sync_in_price(self):
        options = QFileDialog.Options()
        excel_filename, _ = QFileDialog.getOpenFileName(
            self.parent,
            "匯入處方進價檔案",
            '藥品資料.xlsx',
            "xlsx檔案 (*.xlsx)",
            options=options
        )
        if not excel_filename:
            return

        workbook = openpyxl.load_workbook(excel_filename, data_only=True)
        self._sync_medicine(workbook, '單方')
        self._sync_medicine(workbook, '複方')

        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Information)
        msg_box.setWindowTitle('藥品進價檔匯入完成')
        msg_box.setText('''
            <b>藥品進價檔匯入完成.<br>
        ''')
        msg_box.setInformativeText("所有資料均已匯入.")
        msg_box.addButton(QPushButton("確定"), QMessageBox.YesRole)
        msg_box.exec_()

    def _sync_medicine(self, workbook, medicine_type):
        sheet = workbook[medicine_type]

        max_progress = sheet.max_row
        progress_dialog = QtWidgets.QProgressDialog(
            f'正在匯入{medicine_type}資料Excel中, 請稍後...', '取消', 0, max_progress, self
        )

        progress_dialog.setWindowModality(QtCore.Qt.WindowModal)
        progress_dialog.setValue(0)

        start_no = 2
        i = 0
        for row_no in range(start_no, sheet.max_row + start_no+1):
            i += 1
            progress_dialog.setValue(i)

            medicine_name = string_utils.xstr(sheet.cell(row_no, 1).value)
            if medicine_name == '':
                continue

            in_price = string_utils.xstr(sheet.cell(row_no, 4).value)
            try:
                in_price = number_utils.Decimal(in_price)
                in_price = round(in_price, 2)
            except Exception:
                continue

            self.database.exec_sql(f'''
                UPDATE medicine
                SET
                    InPrice = {in_price}
                WHERE
                    MedicineType = "{medicine_type}" AND
                    MedicineName = "{medicine_name}"
            ''')

    # 更改處方類別不自購
    def _no_purchase(self):
        dict_groups_key = self.table_widget_dict_groups.field_value(self.groups_col_no['dict_groups_key'])
        current_row_no = self.ui.tableWidget_dict_groups.currentRow()
        no_purchase = self.table_widget_dict_groups.field_value(self.groups_col_no['dict_groups_no_purchase'])
        if no_purchase == 'Y':
            no_purchase_value = 'NULL'
        else:
            no_purchase_value = '"Y"'

        sql = f'''
            UPDATE dict_groups
            SET
                DictGroupsLevel3 = {no_purchase_value}
            WHERE
                DictGroupsKey = {dict_groups_key}
        '''
        self.database.exec_sql(sql)
        self._read_dict_groups()
        self.ui.tableWidget_dict_groups.setCurrentCell(current_row_no, 1)

    def _sync_drug(self):
        last_dir = system_utils.get_last_directory('處方同步資料')
        options = QFileDialog.Options()
        excel_filename, _ = QFileDialog.getOpenFileName(
            self.parent,
            "同步處方",
            last_dir,  # ← 這裡放入 last_dir
            "xls檔案 (*.xlsx)",
            options=options
        )
        if not excel_filename:
            return

        system_utils.set_last_directory('處方同步資料', excel_filename)
        workbook = openpyxl.load_workbook(excel_filename, data_only=True)
        medicine_type = self.table_widget_dict_groups.field_value(self.groups_col_no['dict_groups_name'])
        self._sync_drug_data(workbook, medicine_type)

        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Information)
        msg_box.setWindowTitle('處方完成')
        msg_box.setText('''
            <b>處方資料同步完成.<br>
        ''')
        msg_box.setInformativeText("所有資料均已同步.")
        msg_box.addButton(QPushButton("確定"), QMessageBox.YesRole)
        msg_box.exec_()

        self.dict_groups_changed()

    def _sync_drug_data(self, workbook, medicine_type):
        sheet = workbook['Sheet1']

        max_progress = sheet.max_row
        progress_dialog = QtWidgets.QProgressDialog(
            f'正在同步{medicine_type}資料Excel中, 請稍後...', '取消', 0, max_progress, self
        )

        progress_dialog.setWindowModality(QtCore.Qt.WindowModal)
        progress_dialog.setValue(0)

        start_no = 2
        i = 0
        for row_no in range(start_no, sheet.max_row + start_no+1):
            i += 1
            progress_dialog.setValue(i)

            medicine_name = string_utils.xstr(sheet.cell(row_no, 5).value)
            if medicine_name == '':
                continue

            medicine_code = string_utils.xstr(sheet.cell(row_no, 3).value)

            input_code = string_utils.xstr(sheet.cell(row_no, 4).value).strip()
            if input_code == '':
                try:
                    import pypinyin
                except Exception:
                    system_utils.pip3_install('pypinyin')

                input_code = string_utils.get_input_code(medicine_name)[:5]

            ins_code = string_utils.xstr(sheet.cell(row_no, 6).value)
            unit = string_utils.xstr(sheet.cell(row_no, 7).value)
            dosage = string_utils.xstr(sheet.cell(row_no, 8).value)
            location = string_utils.xstr(sheet.cell(row_no, 12).value)
            in_price = string_utils.xstr(sheet.cell(row_no, 13).value)
            try:
                in_price = number_utils.Decimal(in_price)
                in_price = round(in_price, 2)
            except Exception:
                in_price = 0

            sale_price = string_utils.xstr(sheet.cell(row_no, 14).value)
            try:
                sale_price = number_utils.Decimal(sale_price)
                sale_price = round(sale_price, 2)
            except Exception:
                sale_price = 0

            self._update_drug_data(
                medicine_type=medicine_type,
                medicine_name=medicine_name,
                medicine_code=medicine_code,
                input_code=input_code,
                ins_code=ins_code,
                unit=unit,
                dosage=dosage,
                location=location,
                in_price=in_price,
                sale_price=sale_price,
            )
            
    def _update_drug_data(self, **kwargs):
        medicine_type = kwargs['medicine_type']
        medicine_name = kwargs['medicine_name']
        sql = f'''
            SELECT MedicineKey FROM medicine
            WHERE
                MedicineType = "{medicine_type}" AND
                MedicineName = "{medicine_name}"
        '''
        rows = self.database.select_record(sql)
        if not rows:
            self._insert_drug_data(**kwargs)
            return

        row = rows[0]

        medicine_key = row['MedicineKey']
        medicine_code = kwargs['medicine_code']
        input_code = kwargs['input_code']
        ins_code = kwargs['ins_code']
        unit = kwargs['unit']
        dosage = kwargs['dosage']
        if dosage == '':
            dosage = 'NULL'

        location = kwargs['location']
        in_price = kwargs['in_price']
        sale_price = kwargs['sale_price']

        sql = f'''
            UPDATE medicine
            SET
                MedicineCode = "{medicine_code}",
                InputCode = "{input_code}",
                InsCode = "{ins_code}",
                Unit = "{unit}",
                Dosage = {dosage},
                Location = "{location}",
                InPrice = {in_price},
                SalePrice = {sale_price}
            WHERE
                MedicineKey = {medicine_key}
        '''
        self.database.exec_sql(sql)

    def _insert_drug_data(self, **kwargs):
        medicine_type = kwargs['medicine_type']
        medicine_code = kwargs['medicine_code']
        medicine_name = kwargs['medicine_name']

        input_code = kwargs['input_code']
        ins_code = kwargs['ins_code']
        unit = kwargs['unit']
        dosage = kwargs['dosage']
        location = kwargs['location']
        in_price = kwargs['in_price']
        sale_price = kwargs['sale_price']

        fields = [
            'MedicineType', 'MedicineCode', 'InputCode', 'MedicineName', 'InsCode', 'Unit',
            'Dosage', 'Location', 'InPrice', 'SalePrice',
        ]
        data = [
            medicine_type, medicine_code, input_code, medicine_name, ins_code, unit,
            dosage, location, in_price, sale_price,
        ]
        self.database.insert_record('medicine', fields, data)
        print('insert:', medicine_type, medicine_name)
