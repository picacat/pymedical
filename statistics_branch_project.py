
# -*- coding: UTF-8 -*-

import subprocess

from openpyxl import Workbook
from openpyxl.styles import Alignment
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QFileDialog

from libs import (db_utils, dialog_utils, module_utils, number_utils,
                  string_utils, system_utils, ui_utils)


# 分院專案統計 2021.02.09
class StatisticsBranchProject(QtWidgets.QMainWindow):
    # 初始化
    def __init__(self, parent=None, *args):
        super(StatisticsBranchProject, self).__init__(parent)
        self.parent = parent
        self.database = args[0]
        self.system_settings = args[1]
        self.ui = None

        self.dialog_setting = {
            "dialog_executed": False,
            "start_date": None,
            "end_date": None,
            "ins_type": None,
            "therapist": None,
        }

        self._set_ui()
        self._set_signal()

    # 解構
    def __del__(self):
        self.close_all()

    # 關閉
    def close_all(self):
        pass

    # 設定GUI
    def _set_ui(self):
        self.ui = ui_utils.load_ui_file(ui_utils.UI_STATISTICS_BRANCH_PROJECT, self)
        system_utils.set_css(self, self.system_settings)
        system_utils.center_window(self)

    # 設定信號
    def _set_signal(self):
        self.ui.action_close.triggered.connect(self.close_form)
        self.ui.action_open_dialog.triggered.connect(self.open_dialog)
        self.ui.action_export_excel.triggered.connect(self._export_to_excel)

    def close_tab(self):
        current_tab = self.parent.ui.tabWidget_window.currentIndex()
        self.parent.close_tab(current_tab)

    def close_form(self):
        self.close_all()
        self.close_tab()

    # 讀取病歷
    def open_dialog(self):
        dialog = dialog_utils.get_dialog_statistics_therapist(
            self, self.database, self.system_settings, '用藥統計', '醫師',
        )

        if self.dialog_setting['dialog_executed']:
            dialog.ui.dateEdit_start_date.setDate(self.dialog_setting['start_date'])
            dialog.ui.dateEdit_end_date.setDate(self.dialog_setting['end_date'])

            if self.dialog_setting['ins_type'] == '全部':
                dialog.ui.radioButton_all.setChecked(True)
            elif self.dialog_setting['ins_type'] == '健保':
                dialog.ui.radioButton_ins.setChecked(True)
            elif self.dialog_setting['ins_type'] == '自費':
                dialog.ui.radioButton_self.setChecked(True)

            dialog.ui.comboBox_therapist.setCurrentText(self.dialog_setting['therapist'])

        if not dialog.exec_():
            dialog.deleteLater()
            return

        start_date = dialog.start_date()
        end_date = dialog.end_date()
        ins_type = dialog.ins_type()
        therapist = dialog.ui.comboBox_therapist.currentText()

        self.dialog_setting['dialog_executed'] = True
        self.dialog_setting['start_date'] = dialog.ui.dateEdit_start_date.date()
        self.dialog_setting['end_date'] = dialog.ui.dateEdit_end_date.date()
        self.dialog_setting['ins_type'] = ins_type
        self.dialog_setting['therapist'] = therapist

        dialog.deleteLater()
        self._set_tab_widget(start_date, end_date, ins_type, therapist)

    def _get_project_items(self, database, start_date, end_date, ins_type, doctor):
        ins_type_condition = f' AND InsType = "{ins_type}"' if ins_type != '全部' else ''
        doctor_type_condition = f' AND Doctor = "{doctor}"' if doctor != '全部' else ''

        sql = f'''
            SELECT prescript.MedicineType, medicine.Project FROM prescript
                LEFT JOIN cases ON prescript.CaseKey = cases.CaseKey
                LEFT JOIN medicine ON medicine.MedicineKey = prescript.MedicineKey
            WHERE
                cases.CaseDate BETWEEN "{start_date}" AND "{end_date}" AND
                prescript.MedicineSet >= 2 AND
                medicine.Project IS NOT NULL
                {ins_type_condition}
                {doctor_type_condition}
            GROUP BY Project
        '''
        rows = database.select_record(sql)

        project_list = []
        for row in rows:
            project_list.append(string_utils.xstr(row['Project']))

        return project_list

    def _get_project_list(self, start_date, end_date, ins_type, doctor):
        project_list = []

        database_list = db_utils.get_host_database_dict(self.database, '分院統計')
        clinic_name = self.system_settings.field('院所名稱')
        database_list[clinic_name] = {'database': self.database}

        for clinic_name in database_list.keys():
            database = database_list[clinic_name]['database']
            project_items = self._get_project_items(database, start_date, end_date, ins_type, doctor)
            for item in project_items:
                if item not in project_list:
                    project_list.append(item)

        return project_list, database_list

    def _set_tab_widget(self, start_date, end_date, ins_type, doctor):
        self.ui.statusbar.showMessage(
            f' 統計期間: 從 {start_date[:10]} 至 {end_date[:10]} 保險: {ins_type} 醫師: {doctor}'
        )

        project_list, database_list = self._get_project_list(start_date, end_date, ins_type, doctor)
        self.ui.tabWidget_statistics_medicine.clear()
        for project_name in project_list:
            self._add_statistic_branch_project_sales(
                database_list, start_date, end_date, ins_type, doctor, project_name
            )

    # 用藥統計內容
    def _add_statistic_branch_project_sales(self, database_list, start_date, end_date, ins_type, doctor, project_name):
        self.tab_statistics_branch_project_sales = module_utils.get_statistics_branch_project_sales(
            self, self.database, database_list, self.system_settings,
            start_date, end_date, ins_type, doctor, project_name,
        )
        self.tab_statistics_branch_project_sales.start_calculate()
        self.ui.tabWidget_statistics_medicine.addTab(self.tab_statistics_branch_project_sales, project_name)

    def _export_to_excel(self):
        start_date = self.dialog_setting['start_date'].toString('yyyy-MM-dd')
        end_date = self.dialog_setting['end_date'].toString('yyyy-MM-dd')
        options = QFileDialog.Options()
        excel_file_name, _ = QFileDialog.getSaveFileName(
            self.parent,
            "匯出分院專案統計",
            f'{start_date}至{end_date}分院專案統計表.xlsx',
            "excel檔案 (*.xlsx);;Text Files (*.txt)", options=options
        )
        if not excel_file_name:
            return

        wb = Workbook()
        for tab_no in range(self.ui.tabWidget_statistics_medicine.count()):
            current_tab = self.ui.tabWidget_statistics_medicine.widget(tab_no)
            table_widget = current_tab.tableWidget_branch_project_sales

            sheet_name = self.ui.tabWidget_statistics_medicine.tabText(tab_no)
            sheet_name = string_utils.remove_illegal_characters(sheet_name)
            sheet_name = sheet_name.replace('/', '-')

            ws = wb.create_sheet(sheet_name, tab_no)
            ws.title = sheet_name
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 8
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 30
            ws.column_dimensions['F'].width = 5
            ws.column_dimensions['G'].width = 5
            ws.column_dimensions['K'].width = 10

            header_row = []
            for col_no in range(table_widget.columnCount()):
                if col_no == 0:  # case_key
                    continue

                header_row.append(table_widget.horizontalHeaderItem(col_no).text())

            ws.append(header_row)

            for row_no in range(table_widget.rowCount()):
                ws.cell(row_no+1, 7).alignment = Alignment(horizontal='center', vertical='center')  # 單位
                ws.cell(row_no+1, 10).alignment = Alignment(horizontal='right', vertical='center')  # 折扣

                row = []
                for col_no in range(table_widget.columnCount()):
                    if col_no == 0:
                        continue

                    item = table_widget.item(row_no, col_no)
                    if item is not None:
                        item_text = item.text()
                    else:
                        item_text = ''

                    if table_widget.item(row_no, 1) is not None and col_no in [3, 6, 8, 9, 12]:  # 非總計欄
                        item_text = number_utils.get_float(item_text)
                    elif col_no in [9, 12]:
                        item_text = number_utils.get_float(item_text)

                    row.append(item_text)

                ws.append(row)

        wb.save(excel_file_name)
        try:
            subprocess.Popen([excel_file_name], shell=True)
        except Exception:
            pass
